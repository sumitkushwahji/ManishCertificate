"""
Config-Based Certificate Generator
===================================
Uses a simple config.json file to specify which files to process.
Perfect for batch processing multiple towers at once.

Usage:
    1. Edit config.json to add your files
    2. Run: python batch_certificate_generator.py
"""

import json
import os
from openpyxl import load_workbook, Workbook
from copy import copy

def load_config(config_file='config.json'):
    """Load configuration from JSON file"""
    # Get base directory (parent folder of this script)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.abspath(os.path.join(script_dir, '..'))
    
    if not os.path.exists(config_file):
        # Create default config
        default_config = {
            "base_directory": base_dir,
            "template_file": "CYBER_PARK_TOWER_A_complete.xlsx",
            "towers": [
                {
                    "name": "Tower B",
                    "input_file": "CP TOWER TowerB CALIBRATION Excel sheet.xlsx",
                    "output_file": "CYBER_PARK_TOWER_B_complete.xlsx",
                    "sheet_prefix": "TowerB"
                },
                {
                    "name": "Tower C",
                    "input_file": "CP TOWER TowerC CALIBRATION Excel - Copy.xlsx",
                    "output_file": "CYBER_PARK_TOWER_C_complete.xlsx",
                    "sheet_prefix": "TowerC"
                },
                {
                    "name": "Ground Floor",
                    "input_file": "CP TOWER_GROUND_FLOOR_SHOP-CALIBRATION Excel - Copy.xlsx",
                    "output_file": "CYBER_PARK_GROUND_FLOOR_SHOP_complete.xlsx",
                    "sheet_prefix": "GF"
                },
                {
                    "name": "Basement",
                    "input_file": "CP TOWER_BASEMENT_CALIBRATION Excel.xlsx",
                    "output_file": "CYBER_PARK_BASEMENT_complete.xlsx",
                    "sheet_prefix": "Basement"
                }
            ]
        }
        with open(config_file, 'w') as f:
            json.dump(default_config, f, indent=4)
        print(f"Created default config file: {config_file}")
        return default_config
    
    with open(config_file, 'r') as f:
        return json.load(f)


def generate_certificates(calibration_file, output_file, sheet_prefix, template_file):
    """Generate certificates from a calibration file"""
    wb_cal = load_workbook(calibration_file)
    ws_cal = wb_cal['Sheet1']
    
    # Extract meter data
    meters = []
    for row in ws_cal.iter_rows(min_row=5, values_only=False):
        if row[0].value and row[1].value:
            before_mwh = row[7].value
            before_kwh = row[8].value
            if before_mwh:
                before_unit, before_value = 'MWH', before_mwh
            elif before_kwh:
                before_unit, before_value = 'KWH', before_kwh
            else:
                before_unit, before_value = None, None
            
            after_mwh = row[13].value if len(row) > 13 else None
            after_kwh = row[14].value if len(row) > 14 else None
            if after_mwh:
                after_unit, after_value = 'MWH', after_mwh
            elif after_kwh:
                after_unit, after_value = 'KWH', after_kwh
            else:
                after_unit, after_value = None, None
            
            meters.append({
                'location': str(row[0].value).strip(),
                'serial': str(row[1].value).strip(),
                'meter_size': row[2].value,
                'before_inlet': row[5].value,
                'before_outlet': row[4].value,
                'before_m3hr': row[6].value,
                'before_unit': before_unit,
                'before_value': before_value,
                'after_inlet': row[11].value if len(row) > 11 else None,
                'after_outlet': row[10].value if len(row) > 10 else None,
                'after_m3hr': row[12].value if len(row) > 12 else None,
                'after_unit': after_unit,
                'after_value': after_value,
            })
    
    # Load template and create new workbook
    wb_template = load_workbook(template_file)
    template_sheet = wb_template[wb_template.sheetnames[0]]
    wb_new = Workbook()
    wb_new.remove(wb_new.active)
    
    # Create sheets
    for meter in meters:
        location_clean = (meter['location'].upper()
                         .replace(' ', '_').replace('(', '').replace(')', '')
                         .replace('&', 'AND').replace('-', '_'))
        sheet_name = f"{sheet_prefix}_{location_clean}"[:31]
        ws_new = wb_new.create_sheet(title=sheet_name)
        
        # Copy template
        for row in template_sheet.iter_rows():
            for cell in row:
                new_cell = ws_new[cell.coordinate]
                if cell.value:
                    new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        for col_letter, col_dim in template_sheet.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = col_dim.width
        for row_num, row_dim in template_sheet.row_dimensions.items():
            ws_new.row_dimensions[row_num].height = row_dim.height
        for merged_cell_range in template_sheet.merged_cells.ranges:
            ws_new.merge_cells(str(merged_cell_range))
        
        # Fill data
        ws_new['B7'].value = f"Serial No: {meter['serial']}"
        ws_new['B8'].value = f"Meter Location : {meter['location']}"
        meter_size = f"DN-{meter['meter_size']}" if meter['meter_size'] else "DN-65"
        ws_new['B9'].value = f"Meter Size : {meter_size}"
        
        if meter['before_unit'] and meter['before_value'] is not None:
            ws_new['I13'].value = f"{meter['before_unit']}= BTU*{meter['before_value']}"
        if meter['before_inlet'] is not None:
            ws_new['D14'].value = float(meter['before_inlet'])
        if meter['before_outlet'] is not None:
            ws_new['D15'].value = float(meter['before_outlet'])
        if meter['before_m3hr'] is not None:
            ws_new['F16'].value = float(meter['before_m3hr'])
        if meter['before_inlet'] and meter['before_outlet']:
            ws_new['D16'].value = abs(float(meter['before_outlet']) - float(meter['before_inlet']))
        
        if meter['after_unit'] and meter['after_value'] is not None:
            ws_new['I19'].value = f"{meter['after_unit']}= BTU*{meter['after_value']}"
        if meter['after_inlet'] is not None:
            ws_new['D20'].value = float(meter['after_inlet'])
        if meter['after_outlet'] is not None:
            ws_new['D21'].value = float(meter['after_outlet'])
        if meter['after_m3hr'] is not None:
            ws_new['F22'].value = float(meter['after_m3hr'])
    
    wb_new.save(output_file)
    wb_template.close()
    wb_cal.close()
    wb_new.close()
    
    return len(meters)


def main():
    """Main batch processing function"""
    print("=" * 70)
    print("  BATCH CERTIFICATE GENERATOR")
    print("=" * 70)
    
    # Load config
    config = load_config()
    base_dir = config['base_directory']
    template_file = os.path.join(base_dir, config['template_file'])
    
    print(f"\nBase Directory: {base_dir}")
    print(f"Template: {config['template_file']}")
    print(f"\nProcessing {len(config['towers'])} tower(s)...\n")
    
    results = []
    for idx, tower in enumerate(config['towers'], 1):
        print(f"[{idx}/{len(config['towers'])}] Processing {tower['name']}...")
        
        input_file = os.path.join(base_dir, tower['input_file'])
        output_file = os.path.join(base_dir, tower['output_file'])
        
        try:
            count = generate_certificates(
                input_file,
                output_file,
                tower['sheet_prefix'],
                template_file
            )
            print(f"     ✓ Created {count} certificates")
            results.append((tower['name'], count, 'SUCCESS'))
        except Exception as e:
            print(f"     ✗ Error: {str(e)}")
            results.append((tower['name'], 0, f'FAILED: {str(e)}'))
    
    # Summary
    print("\n" + "=" * 70)
    print("  BATCH PROCESSING SUMMARY")
    print("=" * 70)
    for name, count, status in results:
        print(f"  {name:20s}: {count:3d} certificates - {status}")
    print("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
