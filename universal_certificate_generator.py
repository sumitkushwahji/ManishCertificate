"""
Universal Certificate Generator
================================
A parameterized script that can generate certificates for ANY calibration file
by simply providing the input file path, output file path, and prefix.

Usage:
    python universal_certificate_generator.py

The script will prompt you for:
1. Path to calibration Excel file
2. Output file name
3. Sheet name prefix (e.g., 'TowerB', 'TowerC', 'GF', 'Basement')
"""

from openpyxl import load_workbook, Workbook
from copy import copy
import os
import sys

def generate_certificates(calibration_file, output_file, sheet_prefix, template_file):
    """
    Generate certificates from a calibration file.
    
    Args:
        calibration_file: Path to the calibration Excel file
        output_file: Path for the output Excel file
        sheet_prefix: Prefix for sheet names (e.g., 'TowerB', 'GF')
        template_file: Path to the template Excel file
    """
    print("=" * 70)
    print(f"Universal Certificate Generator")
    print("=" * 70)
    
    # Step 1: Load calibration data
    print(f"\n[1/5] Loading calibration data from: {os.path.basename(calibration_file)}")
    if not os.path.exists(calibration_file):
        print(f"ERROR: File not found: {calibration_file}")
        return False
    
    wb_cal = load_workbook(calibration_file)
    ws_cal = wb_cal['Sheet1']
    
    # Extract all meter data
    meters = []
    for row in ws_cal.iter_rows(min_row=5, values_only=False):
        if row[0].value and row[1].value:  # Has location and serial
            # Check MWH first, then KWH for Before Calibration
            before_mwh = row[7].value  # Column H
            before_kwh = row[8].value  # Column I
            if before_mwh:
                before_unit = 'MWH'
                before_value = before_mwh
            elif before_kwh:
                before_unit = 'KWH'
                before_value = before_kwh
            else:
                before_unit = None
                before_value = None
            
            # Check MWH first, then KWH for After Calibration
            after_mwh = row[13].value if len(row) > 13 else None  # Column N
            after_kwh = row[14].value if len(row) > 14 else None  # Column O
            if after_mwh:
                after_unit = 'MWH'
                after_value = after_mwh
            elif after_kwh:
                after_unit = 'KWH'
                after_value = after_kwh
            else:
                after_unit = None
                after_value = None
            
            meter_data = {
                'location': str(row[0].value).strip(),
                'serial': str(row[1].value).strip(),
                'meter_size': row[2].value,
                'before_inlet': row[5].value,      # Column F (Inlet Temp)
                'before_outlet': row[4].value,     # Column E (Outlet Temp)
                'before_m3hr': row[6].value,       # Column G
                'before_unit': before_unit,
                'before_value': before_value,
                'after_inlet': row[11].value if len(row) > 11 else None,    # Column L
                'after_outlet': row[10].value if len(row) > 10 else None,   # Column K
                'after_m3hr': row[12].value if len(row) > 12 else None,     # Column M
                'after_unit': after_unit,
                'after_value': after_value,
            }
            meters.append(meter_data)
    
    print(f"   ✓ Found {len(meters)} meters")
    
    # Step 2: Load template
    print(f"\n[2/5] Loading template from: {os.path.basename(template_file)}")
    if not os.path.exists(template_file):
        print(f"ERROR: Template file not found: {template_file}")
        return False
    
    wb_template = load_workbook(template_file)
    template_sheet = wb_template[wb_template.sheetnames[0]]
    print(f"   ✓ Template loaded")
    
    # Step 3: Create new workbook
    print(f"\n[3/5] Creating new workbook...")
    wb_new = Workbook()
    wb_new.remove(wb_new.active)  # Remove default sheet
    
    # Step 4: Create certificate sheets
    print(f"\n[4/5] Creating certificate sheets...")
    for idx, meter in enumerate(meters, 1):
        # Create sheet name from location
        location_clean = (meter['location'].upper()
                         .replace(' ', '_')
                         .replace('(', '')
                         .replace(')', '')
                         .replace('&', 'AND')
                         .replace('-', '_'))
        sheet_name = f"{sheet_prefix}_{location_clean}"[:31]  # Excel limit is 31 chars
        
        print(f"   [{idx}/{len(meters)}] {sheet_name}")
        
        # Create new sheet by copying template structure
        ws_new = wb_new.create_sheet(title=sheet_name)
        
        # Copy all cells from template
        for row in template_sheet.iter_rows():
            for cell in row:
                new_cell = ws_new[cell.coordinate]
                if cell.value:
                    new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # Copy column widths and row heights
        for col_letter, col_dim in template_sheet.column_dimensions.items():
            ws_new.column_dimensions[col_letter].width = col_dim.width
        for row_num, row_dim in template_sheet.row_dimensions.items():
            ws_new.row_dimensions[row_num].height = row_dim.height
        
        # Copy merged cells
        for merged_cell_range in template_sheet.merged_cells.ranges:
            ws_new.merge_cells(str(merged_cell_range))
        
        # Fill in the meter data
        ws_new['B7'].value = f"Serial No: {meter['serial']}"
        ws_new['B8'].value = f"Meter Location : {meter['location']}"
        meter_size_text = f"DN-{meter['meter_size']}" if meter['meter_size'] else "DN-65"
        ws_new['B9'].value = f"Meter Size : {meter_size_text}"
        
        # Before Calibration
        if meter['before_unit'] and meter['before_value'] is not None:
            ws_new['I13'].value = f"{meter['before_unit']}= BTU*{meter['before_value']}"
        if meter['before_inlet'] is not None:
            ws_new['D14'].value = float(meter['before_inlet'])
        if meter['before_outlet'] is not None:
            ws_new['D15'].value = float(meter['before_outlet'])
        if meter['before_m3hr'] is not None:
            ws_new['F16'].value = float(meter['before_m3hr'])
        if meter['before_inlet'] is not None and meter['before_outlet'] is not None:
            delta_t = abs(float(meter['before_outlet']) - float(meter['before_inlet']))
            ws_new['D16'].value = delta_t
        
        # After Calibration
        if meter['after_unit'] and meter['after_value'] is not None:
            ws_new['I19'].value = f"{meter['after_unit']}= BTU*{meter['after_value']}"
        if meter['after_inlet'] is not None:
            ws_new['D20'].value = float(meter['after_inlet'])
        if meter['after_outlet'] is not None:
            ws_new['D21'].value = float(meter['after_outlet'])
        if meter['after_m3hr'] is not None:
            ws_new['F22'].value = float(meter['after_m3hr'])
    
    # Step 5: Save the file
    print(f"\n[5/5] Saving certificate file...")
    wb_new.save(output_file)
    print(f"   ✓ File saved: {output_file}")
    
    # Close workbooks
    wb_template.close()
    wb_cal.close()
    wb_new.close()
    
    print("\n" + "=" * 70)
    print(f"✓ SUCCESS! Created {len(meters)} certificate sheets")
    print(f"✓ Output: {output_file}")
    print("=" * 70)
    
    return True


def main():
    """Main function with interactive prompts"""
    print("\n" + "=" * 70)
    print("  UNIVERSAL CERTIFICATE GENERATOR")
    print("=" * 70)
    
    # Get base directory (parent folder)
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    
    # Interactive mode
    print("\nEnter the required information:")
    print("-" * 70)
    
    # Get calibration file
    cal_file = input("\n1. Calibration file name (e.g., 'CP TOWER B CALIBRATION.xlsx')\n   > ").strip()
    if not cal_file:
        print("ERROR: File name is required")
        return
    
    cal_file_path = os.path.join(base_dir, cal_file)
    
    # Get output file
    output_file = input("\n2. Output file name (e.g., 'CYBER_PARK_TOWER_B_complete.xlsx')\n   > ").strip()
    if not output_file:
        print("ERROR: Output file name is required")
        return
    
    output_file_path = os.path.join(base_dir, output_file)
    
    # Get sheet prefix
    sheet_prefix = input("\n3. Sheet name prefix (e.g., 'TowerB', 'GF', 'Basement')\n   > ").strip()
    if not sheet_prefix:
        print("ERROR: Sheet prefix is required")
        return
    
    # Template file (fixed)
    template_file = os.path.join(base_dir, 'CYBER_PARK_TOWER_A_complete.xlsx')
    
    print("\n" + "-" * 70)
    print("Configuration:")
    print(f"  Input:    {os.path.basename(cal_file_path)}")
    print(f"  Output:   {os.path.basename(output_file_path)}")
    print(f"  Prefix:   {sheet_prefix}")
    print(f"  Template: {os.path.basename(template_file)}")
    print("-" * 70)
    
    confirm = input("\nProceed with generation? (yes/no): ").strip().lower()
    if confirm not in ['yes', 'y']:
        print("Cancelled.")
        return
    
    # Generate certificates
    success = generate_certificates(cal_file_path, output_file_path, sheet_prefix, template_file)
    
    if success:
        print("\n✓ Certificate generation completed successfully!")
    else:
        print("\n✗ Certificate generation failed. Please check the errors above.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
