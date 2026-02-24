"""
GUI Certificate Generator with PDF Export
==========================================
User-friendly graphical interface for generating certificates and exporting to PDF.

Features:
- File browser for easy file selection
- Real-time progress display
- Success/error notifications
- PDF Export: Export one, multiple, or all certificate sheets to PDF
- No command-line knowledge required

Usage:
    python gui_certificate_generator.py
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Listbox, Scrollbar, MULTIPLE
import os
from openpyxl import load_workbook, Workbook
from copy import copy
import threading


class CertificateGeneratorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Certificate Generator & PDF Export")
        self.root.geometry("850x700")
        self.root.resizable(False, False)
        
        # Base directory (parent folder)
        self.base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        
        # Create UI
        self.create_widgets()
    
    def create_widgets(self):
        """Create all UI widgets"""
        # Title
        title = tk.Label(self.root, text="📄 Certificate Generator & PDF Export", 
                        font=("Arial", 18, "bold"), pady=15)
        title.pack()
        
        # Create notebook (tabbed interface)
        notebook = ttk.Notebook(self.root)
        notebook.pack(expand=True, fill="both", padx=20, pady=10)
        
        # Tab 1: Certificate Generation
        tab_generate = tk.Frame(notebook)
        notebook.add(tab_generate, text="📝 Generate Certificates")
        
        # Tab 2: PDF Export
        tab_pdf = tk.Frame(notebook)
        notebook.add(tab_pdf, text="📄 Export to PDF")
        
        # Create widgets for each tab
        self.create_generate_tab(tab_generate)
        self.create_pdf_tab(tab_pdf)
    
    def create_generate_tab(self, parent):
        """Create widgets for certificate generation tab"""
        # Main frame
        main_frame = tk.Frame(parent, padx=30, pady=10)
        main_frame.pack(expand=True, fill="both")
        
        # Calibration file selection
        tk.Label(main_frame, text="1. Calibration File:", 
                font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="w", pady=5)
        
        file_frame = tk.Frame(main_frame)
        file_frame.grid(row=1, column=0, sticky="ew", pady=5)
        
        self.file_entry = tk.Entry(file_frame, width=60)
        self.file_entry.pack(side="left", padx=(0, 10))
        
        browse_btn = tk.Button(file_frame, text="Browse", command=self.browse_file)
        browse_btn.pack(side="left")
        
        # Output folder selection
        tk.Label(main_frame, text="2. Output Folder:", 
                font=("Arial", 11, "bold")).grid(row=2, column=0, sticky="w", pady=(20, 5))
        
        folder_frame = tk.Frame(main_frame)
        folder_frame.grid(row=3, column=0, sticky="ew", pady=5)
        
        self.output_folder_entry = tk.Entry(folder_frame, width=60)
        self.output_folder_entry.pack(side="left", padx=(0, 10))
        
        # Set default output folder
        default_output_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Output')
        self.output_folder_entry.insert(0, default_output_folder)
        
        browse_folder_btn = tk.Button(folder_frame, text="Browse", command=self.browse_output_folder)
        browse_folder_btn.pack(side="left")
        
        # Output file name
        tk.Label(main_frame, text="3. Output File Name:", 
                font=("Arial", 11, "bold")).grid(row=4, column=0, sticky="w", pady=(20, 5))
        
        self.output_entry = tk.Entry(main_frame, width=60)
        self.output_entry.grid(row=5, column=0, sticky="w", pady=5)
        self.output_entry.insert(0, "CYBER_PARK_OUTPUT_complete.xlsx")
        
        # Sheet prefix
        tk.Label(main_frame, text="4. Sheet Name Prefix:", 
                font=("Arial", 11, "bold")).grid(row=6, column=0, sticky="w", pady=(20, 5))
        
        self.prefix_entry = tk.Entry(main_frame, width=30)
        self.prefix_entry.grid(row=7, column=0, sticky="w", pady=5)
        self.prefix_entry.insert(0, "Tower")
        
        # Generate button
        self.generate_btn = tk.Button(main_frame, text="🚀 Generate Certificates", 
                                     command=self.generate_certificates,
                                     font=("Arial", 12, "bold"),
                                     bg="#4CAF50", fg="white",
                                     padx=20, pady=10,
                                     cursor="hand2")
        self.generate_btn.grid(row=8, column=0, pady=20)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, length=500, mode='indeterminate')
        self.progress.grid(row=9, column=0, pady=10)
        
        # Status label
        self.status_label = tk.Label(main_frame, text="Ready", 
                                     font=("Arial", 10), fg="gray")
        self.status_label.grid(row=10, column=0, pady=5)
    
    def create_pdf_tab(self, parent):
        """Create widgets for PDF export tab"""
        # Main frame
        main_frame = tk.Frame(parent, padx=30, pady=20)
        main_frame.pack(expand=True, fill="both")
        
        # Excel file selection
        tk.Label(main_frame, text="1. Select Certificate Excel File:", 
                font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="w", pady=5)
        
        file_frame = tk.Frame(main_frame)
        file_frame.grid(row=1, column=0, sticky="ew", pady=5)
        
        self.pdf_file_entry = tk.Entry(file_frame, width=65)
        self.pdf_file_entry.pack(side="left", padx=(0, 10))
        
        browse_pdf_btn = tk.Button(file_frame, text="Browse", command=self.browse_pdf_file)
        browse_pdf_btn.pack(side="left")
        
        # Sheet selection
        tk.Label(main_frame, text="2. Select Sheets to Export:", 
                font=("Arial", 11, "bold")).grid(row=2, column=0, sticky="w", pady=(20, 5))
        
        # Listbox with scrollbar
        list_frame = tk.Frame(main_frame)
        list_frame.grid(row=3, column=0, sticky="ew", pady=5)
        
        scrollbar = Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.sheets_listbox = Listbox(list_frame, selectmode=MULTIPLE, height=8, 
                                      yscrollcommand=scrollbar.set, width=70)
        self.sheets_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.sheets_listbox.yview)
        
        # Selection buttons
        btn_frame = tk.Frame(main_frame)
        btn_frame.grid(row=4, column=0, pady=10)
        
        tk.Button(btn_frame, text="Select All", command=self.select_all_sheets,
                 bg="#2196F3", fg="white", padx=15, pady=5).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Clear Selection", command=self.clear_selection,
                 bg="#FF9800", fg="white", padx=15, pady=5).pack(side="left", padx=5)
        
        # Output folder selection
        tk.Label(main_frame, text="3. Output Folder:", 
                font=("Arial", 11, "bold")).grid(row=5, column=0, sticky="w", pady=(20, 5))
        
        folder_frame = tk.Frame(main_frame)
        folder_frame.grid(row=6, column=0, sticky="ew", pady=5)
        
        self.pdf_output_entry = tk.Entry(folder_frame, width=65)
        self.pdf_output_entry.pack(side="left", padx=(0, 10))
        self.pdf_output_entry.insert(0, os.path.join(self.base_dir, "PDF_Certificates"))
        
        browse_folder_btn = tk.Button(folder_frame, text="Browse", command=self.browse_output_folder)
        browse_folder_btn.pack(side="left")
        
        # Export button
        self.export_btn = tk.Button(main_frame, text="📄 Export to PDF", 
                                    command=self.export_to_pdf,
                                    font=("Arial", 12, "bold"),
                                    bg="#E91E63", fg="white",
                                    padx=20, pady=10,
                                    cursor="hand2")
        self.export_btn.grid(row=7, column=0, pady=20)
        
        # Progress bar
        self.pdf_progress = ttk.Progressbar(main_frame, length=500, mode='determinate')
        self.pdf_progress.grid(row=8, column=0, pady=10)
        
        # Status label
        self.pdf_status_label = tk.Label(main_frame, text="Select a certificate file to begin", 
                                        font=("Arial", 10), fg="gray")
        self.pdf_status_label.grid(row=9, column=0, pady=5)
    
    def browse_file(self):
        """Open file browser dialog"""
        filename = filedialog.askopenfilename(
            initialdir=self.base_dir,
            title="Select Calibration File",
            filetypes=(("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*"))
        )
        if filename:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, filename)
            
            # Auto-generate output name based on input
            base_name = os.path.splitext(os.path.basename(filename))[0]
            # Extract tower name (e.g., "TowerB", "GF", "Basement")
            if "TowerB" in base_name or "TOWER B" in base_name.upper():
                prefix = "TowerB"
            elif "TowerC" in base_name or "TOWER C" in base_name.upper():
                prefix = "TowerC"
            elif "GROUND" in base_name.upper():
                prefix = "GF"
            elif "BASEMENT" in base_name.upper():
                prefix = "Basement"
            else:
                prefix = "Tower"
            
            self.prefix_entry.delete(0, tk.END)
            self.prefix_entry.insert(0, prefix)
            
            output_name = f"CYBER_PARK_{prefix.upper()}_complete.xlsx"
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, output_name)
    
    def browse_output_folder(self):
        """Open folder browser for output folder"""
        folder = filedialog.askdirectory(
            initialdir=self.output_folder_entry.get() if self.output_folder_entry.get() else os.path.dirname(os.path.abspath(__file__)),
            title="Select Output Folder"
        )
        if folder:
            self.output_folder_entry.delete(0, tk.END)
            self.output_folder_entry.insert(0, folder)
    
    def update_status(self, message, color="black"):
        """Update status label"""
        self.status_label.config(text=message, fg=color)
        self.root.update()
    
    def generate_certificates(self):
        """Generate certificates in a separate thread"""
        calibration_file = self.file_entry.get().strip()
        output_folder = self.output_folder_entry.get().strip()
        output_file = self.output_entry.get().strip()
        sheet_prefix = self.prefix_entry.get().strip()
        
        # Validation
        if not calibration_file:
            messagebox.showerror("Error", "Please select a calibration file")
            return
        
        if not os.path.exists(calibration_file):
            messagebox.showerror("Error", f"File not found: {calibration_file}")
            return
        
        if not output_folder:
            messagebox.showerror("Error", "Please select an output folder")
            return
        
        # Create output folder if it doesn't exist
        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
            except Exception as e:
                messagebox.showerror("Error", f"Cannot create output folder:\n{e}")
                return
        
        if not output_file:
            messagebox.showerror("Error", "Please enter an output file name")
            return
        
        if not sheet_prefix:
            messagebox.showerror("Error", "Please enter a sheet prefix")
            return
        
        # Disable button and start progress
        self.generate_btn.config(state="disabled")
        self.progress.config(mode='determinate', value=0, maximum=100)
        self.update_status("⏳ Generating certificates...", "blue")
        
        # Run in thread to avoid blocking UI
        thread = threading.Thread(target=self._generate_worker, 
                                 args=(calibration_file, output_folder, output_file, sheet_prefix))
        thread.start()
    
    def _generate_worker(self, calibration_file, output_folder, output_file, sheet_prefix):
        """Worker function for certificate generation"""
        try:
            # Find template file in Base folder
            base_folder = os.path.join(os.path.dirname(__file__), 'Base')
            template_files = [f for f in os.listdir(base_folder) if f.endswith('.xlsx') and not f.startswith('~$')]
            
            if not template_files:
                self.progress.config(value=0)
                self.generate_btn.config(state="normal")
                self.update_status("✗ No template file found", "red")
                messagebox.showerror("Template Missing", 
                                   f"No Excel template file found in:\n{base_folder}\n\n"
                                   f"Please place your certificate template (.xlsx) in the Base folder.")
                return
            
            template_file = os.path.join(base_folder, template_files[0])
            print(f"DEBUG: Using template file: {template_file}")
            
            # Save output in the selected output folder
            output_path = os.path.join(output_folder, output_file)
            print(f"DEBUG: Output will be saved at: {output_path}")
            
            # Check if output file is already open
            if os.path.exists(output_path):
                try:
                    # Try to open in write mode to check if file is locked
                    with open(output_path, 'a'):
                        pass
                except PermissionError:
                    self.progress.config(value=0)
                    self.generate_btn.config(state="normal")
                    self.update_status(f"✗ File is open in Excel", "red")
                    messagebox.showerror("File In Use", 
                                       f"The output file is currently open:\n{output_file}\n\n"
                                       f"Please close it in Excel and try again.")
                    return
            
            # Generate certificates with progress callback
            def progress_callback(current, total):
                self.root.after(0, lambda: self.update_status(f"⏳ Creating certificate {current} of {total}...", "blue"))
                self.root.after(0, lambda: self.progress.config(value=current, maximum=total))
            
            count = self._generate(calibration_file, output_path, sheet_prefix, template_file, progress_callback)
            
            # Success
            self.progress.config(value=0)
            self.generate_btn.config(state="normal")
            self.update_status(f"✓ Success! Created {count} certificates", "green")
            
            # Show success message with option to open folder
            result = messagebox.askyesno("Success", 
                              f"Successfully created {count} certificate sheets!\n\n"
                              f"Output file saved at:\n{output_path}\n\n"
                              f"Do you want to open the folder?")
            
            if result:
                # Open folder and select the file
                import subprocess
                subprocess.Popen(f'explorer /select,"{output_path}"')
        
        except PermissionError as e:
            # File access error
            self.progress.config(value=0)
            self.generate_btn.config(state="normal")
            self.update_status(f"✗ File is locked", "red")
            messagebox.showerror("File Access Error", 
                               f"Cannot access the file (it may be open in Excel):\n\n"
                               f"{str(e)}\n\n"
                               f"Please close the file and try again.")
        
        except Exception as e:
            # Other errors
            self.root.after(0, lambda: self.progress.config(value=0))
            self.root.after(0, lambda: self.generate_btn.config(state="normal"))
            self.root.after(0, lambda: self.update_status(f"✗ Error: {str(e)}", "red"))
            self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to generate certificates:\n\n{str(e)}"))
    
    def _generate(self, calibration_file, output_file, sheet_prefix, template_file, progress_callback=None):
        """Core generation logic using Excel COM to preserve images"""
        import win32com.client
        
        # Step 1: Extract meter data
        wb_cal = load_workbook(calibration_file)
        ws_cal = wb_cal['Sheet1']
        
        meters = []
        for row in ws_cal.iter_rows(min_row=5, values_only=False):
            if row[0].value and row[1].value:
                before_mwh = row[7].value
                before_kwh = row[8].value
                if before_mwh:
                    before_unit, before_value = 'MWH', before_mwh
                elif before_kwh:
                    before_unit, before_value = 'KWH', before_kwh
                else:
                    before_unit, before_value = None, None
                
                after_mwh = row[13].value if len(row) > 13 else None
                after_kwh = row[14].value if len(row) > 14 else None
                if after_mwh:
                    after_unit, after_value = 'MWH', after_mwh
                elif after_kwh:
                    after_unit, after_value = 'KWH', after_kwh
                else:
                    after_unit, after_value = None, None
                
                meters.append({
                    'location': str(row[0].value).strip(),
                    'serial': str(row[1].value).strip(),
                    'meter_size': row[2].value,
                    'before_inlet': row[5].value,
                    'before_outlet': row[4].value,
                    'before_m3hr': row[6].value,
                    'before_unit': before_unit,
                    'before_value': before_value,
                    'after_inlet': row[11].value if len(row) > 11 else None,
                    'after_outlet': row[10].value if len(row) > 10 else None,
                    'after_m3hr': row[12].value if len(row) > 12 else None,
                    'after_unit': after_unit,
                    'after_value': after_value,
                })
        
        wb_cal.close()
        
        # Step 2: Copy template to output
        output_path = os.path.abspath(output_file)
        template_path = os.path.abspath(template_file)
        
        # Try to remove existing file
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except PermissionError:
                raise PermissionError(f"Cannot delete existing file (it may be open in Excel): {output_path}")
        
        # Step 3: Copy template file to output location
        import shutil
        shutil.copy2(template_path, output_path)
        print(f"DEBUG: Copied template to {output_path}")
        
        # Step 4: Use Excel COM to duplicate sheets
        excel = None
        wb_new = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open the copied file (not creating new workbook)
            wb_new = excel.Workbooks.Open(output_path)
            template_ws = wb_new.Worksheets(1)
            
            print(f"DEBUG: Starting with {wb_new.Worksheets.Count} sheet(s)")
            print(f"DEBUG: Processing {len(meters)} meters")
            
            # Step 4: Create certificate sheets
            for idx, meter in enumerate(meters, 1):
                # Update progress
                if progress_callback:
                    progress_callback(idx, len(meters))
                
                print(f"DEBUG: Processing meter {idx}/{len(meters)}: {meter['location']}")
                
                # Sanitize location name for Excel sheet name
                # Excel doesn't allow: : \ / ? * [ ]
                location_clean = (meter['location'].upper()
                                 .replace(' ', '_')
                                 .replace('(', '').replace(')', '')
                                 .replace('&', 'AND')
                                 .replace('-', '_')
                                 .replace(':', '')
                                 .replace('\\', '')
                                 .replace('/', '')
                                 .replace('?', '')
                                 .replace('*', '')
                                 .replace('[', '')
                                 .replace(']', '')
                                 .replace("'", '')
                                 .replace('"', ''))
                
                # Ensure name doesn't exceed 31 characters and isn't empty
                sheet_name = f"{sheet_prefix}_{location_clean}"[:31].strip('_')
                if not sheet_name:
                    sheet_name = f"{sheet_prefix}_Sheet{idx}"
                
                print(f"DEBUG: Sheet name: {sheet_name}")
                
                # Copy template sheet within the same workbook
                # First iteration uses existing sheet, subsequent iterations create copies
                if idx == 1:
                    # Use the existing first sheet
                    ws_new = wb_new.Worksheets(1)
                    ws_new.Name = sheet_name
                else:
                    # Copy the template sheet (correct syntax: Before=None, After=target_sheet)
                    print(f"DEBUG: Before copy - Count: {wb_new.Worksheets.Count}")
                    template_ws.Copy(None, wb_new.Worksheets(wb_new.Worksheets.Count))
                    print(f"DEBUG: After copy - Count: {wb_new.Worksheets.Count}")
                    ws_new = wb_new.Worksheets(wb_new.Worksheets.Count)
                    ws_new.Name = sheet_name
                
                print(f"DEBUG: After processing sheet {idx}, workbook has {wb_new.Worksheets.Count} sheets")
                
                # Fill data
                ws_new.Range("B7").Value = f"Serial No: {meter['serial']}"
                ws_new.Range("B8").Value = f"Meter Location : {meter['location']}"
                meter_size = f"DN-{meter['meter_size']}" if meter['meter_size'] else "DN-65"
                ws_new.Range("B9").Value = f"Meter Size : {meter_size}"
                
                if meter['before_unit'] and meter['before_value'] is not None:
                    ws_new.Range("I13").Value = f"{meter['before_unit']}= BTU*{meter['before_value']}"
                if meter['before_inlet'] is not None:
                    ws_new.Range("D14").Value = float(meter['before_inlet'])
                if meter['before_outlet'] is not None:
                    ws_new.Range("D15").Value = float(meter['before_outlet'])
                if meter['before_m3hr'] is not None:
                    ws_new.Range("F16").Value = float(meter['before_m3hr'])
                if meter['before_inlet'] and meter['before_outlet']:
                    ws_new.Range("D16").Value = abs(float(meter['before_outlet']) - float(meter['before_inlet']))
                
                if meter['after_unit'] and meter['after_value'] is not None:
                    ws_new.Range("I19").Value = f"{meter['after_unit']}= BTU*{meter['after_value']}"
                if meter['after_inlet'] is not None:
                    ws_new.Range("D20").Value = float(meter['after_inlet'])
                if meter['after_outlet'] is not None:
                    ws_new.Range("D21").Value = float(meter['after_outlet'])
                if meter['after_m3hr'] is not None:
                    ws_new.Range("F22").Value = float(meter['after_m3hr'])
            
            print(f"DEBUG: Loop complete. Workbook has {wb_new.Worksheets.Count} sheets")
            
            # No need to delete default sheets - we're working with copied file
            
            print(f"DEBUG: After cleanup, workbook has {wb_new.Worksheets.Count} sheets")
            
            # Save and close
            print(f"DEBUG: Saving workbook...")
            wb_new.Save()  # Use Save() instead of SaveAs() since file already exists
            print(f"DEBUG: Closing workbook...")
            wb_new.Close(SaveChanges=False)
            print(f"DEBUG: Quitting Excel...")
            excel.Quit()
            print(f"DEBUG: Done! Created {len(meters)} certificates")
            
            return len(meters)
            
        finally:
            # Ensure Excel is closed even if there's an error
            try:
                if wb_new:
                    wb_new.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
            except:
                pass
    
    def browse_pdf_file(self):
        """Open file browser for PDF source Excel file"""
        filename = filedialog.askopenfilename(
            initialdir=self.base_dir,
            title="Select Certificate Excel File",
            filetypes=(("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*"))
        )
        if filename:
            self.pdf_file_entry.delete(0, tk.END)
            self.pdf_file_entry.insert(0, filename)
            self.load_sheets(filename)
    
    def browse_output_folder(self):
        """Open folder browser for PDF output location"""
        folder = filedialog.askdirectory(
            initialdir=self.base_dir,
            title="Select Output Folder for PDFs"
        )
        if folder:
            self.pdf_output_entry.delete(0, tk.END)
            self.pdf_output_entry.insert(0, folder)
    
    def load_sheets(self, excel_file):
        """Load sheet names from Excel file into listbox"""
        try:
            wb = load_workbook(excel_file, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            # Clear and populate listbox
            self.sheets_listbox.delete(0, tk.END)
            for sheet in sheet_names:
                self.sheets_listbox.insert(tk.END, sheet)
            
            self.update_pdf_status(f"Loaded {len(sheet_names)} sheet(s)", "green")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheets:\n\n{str(e)}")
            self.update_pdf_status(f"Error loading file", "red")
    
    def select_all_sheets(self):
        """Select all sheets in listbox"""
        self.sheets_listbox.select_set(0, tk.END)
    
    def clear_selection(self):
        """Clear all selections in listbox"""
        self.sheets_listbox.selection_clear(0, tk.END)
    
    def update_pdf_status(self, message, color="black"):
        """Update PDF status label"""
        self.pdf_status_label.config(text=message, fg=color)
        self.root.update()
    
    def export_to_pdf(self):
        """Export selected sheets to PDF"""
        excel_file = self.pdf_file_entry.get().strip()
        output_folder = self.pdf_output_entry.get().strip()
        selected_indices = self.sheets_listbox.curselection()
        
        # Validation
        if not excel_file:
            messagebox.showerror("Error", "Please select a certificate Excel file")
            return
        
        if not os.path.exists(excel_file):
            messagebox.showerror("Error", f"File not found: {excel_file}")
            return
        
        if not selected_indices:
            messagebox.showerror("Error", "Please select at least one sheet to export")
            return
        
        if not output_folder:
            messagebox.showerror("Error", "Please select an output folder")
            return
        
        # Get selected sheet names
        selected_sheets = [self.sheets_listbox.get(i) for i in selected_indices]
        
        # Disable button and start export
        self.export_btn.config(state="disabled")
        self.pdf_progress['value'] = 0
        self.update_pdf_status(f"Exporting {len(selected_sheets)} sheet(s) to PDF...", "blue")
        
        # Run in thread to avoid blocking UI
        thread = threading.Thread(target=self._export_worker, 
                                 args=(excel_file, output_folder, selected_sheets))
        thread.start()
    
    def _export_worker(self, excel_file, output_folder, selected_sheets):
        """Worker function for PDF export"""
        try:
            # Create output folder if it doesn't exist
            os.makedirs(output_folder, exist_ok=True)
            
            # Import win32com for Excel automation
            try:
                import win32com.client
            except ImportError:
                self.root.after(0, lambda: messagebox.showerror("Error", 
                    "PDF export requires pywin32 package.\n\n"
                    "Install it with:\n"
                    ".venv\\Scripts\\pip.exe install pywin32"))
                self.export_btn.config(state="normal")
                self.update_pdf_status("✗ Missing pywin32 package", "red")
                return
            
            # Initialize Excel application
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open workbook
            abs_path = os.path.abspath(excel_file)
            wb = excel.Workbooks.Open(abs_path)
            
            total = len(selected_sheets)
            exported = []
            failed = []
            
            for idx, sheet_name in enumerate(selected_sheets, 1):
                try:
                    # Update progress
                    progress = (idx / total) * 100
                    self.pdf_progress['value'] = progress
                    self.update_pdf_status(f"Exporting {idx}/{total}: {sheet_name}", "blue")
                    
                    # Get worksheet
                    ws = wb.Worksheets(sheet_name)
                    
                    # Create PDF filename
                    pdf_filename = f"{sheet_name}.pdf"
                    pdf_path = os.path.join(output_folder, pdf_filename)
                    
                    # Export to PDF
                    ws.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF
                    exported.append(sheet_name)
                    
                except Exception as e:
                    failed.append((sheet_name, str(e)))
            
            # Close workbook and Excel
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            # Success
            self.export_btn.config(state="normal")
            self.pdf_progress['value'] = 100
            
            if failed:
                fail_msg = "\n".join([f"- {name}: {err}" for name, err in failed])
                self.update_pdf_status(f"✓ Exported {len(exported)}, {len(failed)} failed", "orange")
                messagebox.showwarning("Partial Success", 
                    f"Exported {len(exported)} PDF(s) successfully.\n\n"
                    f"Failed ({len(failed)}):\n{fail_msg}\n\n"
                    f"Output folder: {output_folder}")
            else:
                self.update_pdf_status(f"✓ Successfully exported {len(exported)} PDF(s)", "green")
                messagebox.showinfo("Success", 
                    f"Successfully exported {len(exported)} PDF file(s)!\n\n"
                    f"Output folder: {output_folder}")
        
        except Exception as e:
            # Error
            self.export_btn.config(state="normal")
            self.pdf_progress['value'] = 0
            self.update_pdf_status(f"✗ Export failed", "red")
            messagebox.showerror("Error", f"Failed to export PDFs:\n\n{str(e)}")


def main():
    """Launch the GUI application"""
    root = tk.Tk()
    app = CertificateGeneratorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
